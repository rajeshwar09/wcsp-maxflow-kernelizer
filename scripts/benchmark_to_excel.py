#!/usr/bin/python3
"""
Runs a kernelizer harness N times on each .wcsp file, records every per-stage
timing, and writes an Excel workbook with two sheets:

  "Raw Runs"  : every individual run's timing for every file/stage
  "Averages"  : mean of the N runs per file per stage, plus min/median/max/spread for the kernelize stage so measurement stability is visible

This script produces TABLES ONLY. Written analysis is not generated here.

Every run writes two files:
  <out>.xlsx           the workbook
  <out>.xlsx.raw.json  the raw numbers, written before the workbook so a long benchmark is never lost if Excel writing fails

usage

# 1) benchmark one or more .wcsp files, 5 repeats each
python3 scripts/benchmark_to_excel.py \\
	--harness ./e2e_pipeline_test \\
	--kernelizer "CPU max-flow" \\
	--repeats 5 \\
	--out results/maxflow_06.xlsx \\
	data/wcsp/06_1M.wcsp

# 2) combine several .raw.json files into one workbook + one combined .raw.json
python3 scripts/benchmark_to_excel.py --merge-raw \\
	results/maxflow_small.xlsx.raw.json results/maxflow_06.xlsx.raw.json \\
	--out results/maxflow_all.xlsx

# 3) compare two kernelizers side by side (takes the two combined .raw.json files)
python3 scripts/benchmark_to_excel.py --compare \\
	results/maxflow_all.xlsx.raw.json results/gurobi_all.xlsx.raw.json \\
	--out results/comparison.xlsx

# 4) rebuild a workbook from a .raw.json without re-running the benchmark
python3 scripts/benchmark_to_excel.py \\
	--from-json results/maxflow_all.xlsx.raw.json \\
	--out results/maxflow_all.xlsx

Expected harness stdout (every harness must print these lines):
[parse DIMACS]        <sec> s
[toPolynomial all]    <sec> s
[ccg.addPolynomial]   <sec> s
[ccg.simplify]        <sec> s
[getGraph copy]       <sec> s
[<KERNELIZER LABEL>]  <sec> s      # e.g. [KernelizerMaxflow] or [KernelizerLP]
TOTAL: <sec> s
Remnant s=<val>, resolved=<int>
"""
import argparse
import glob
import json
import os
import re
import statistics
import subprocess
import sys

# Stage keys in order. The "kernelize" stage label varies by kernelizer, so I match it with a regex other than a fixed string
STAGES = ["parse", "toPoly", "addPoly", "simplify", "getGraph", "kernelize", "total"]
STAGE_PATTERNS = {
	"parse":     r"\[parse DIMACS\]\s+([\d.eE+-]+) s",
	"toPoly":    r"\[toPolynomial all\]\s+([\d.eE+-]+) s",
	"addPoly":   r"\[ccg\.addPolynomial\]\s+([\d.eE+-]+) s",
	"simplify":  r"\[ccg\.simplify\]\s+([\d.eE+-]+) s",
	"getGraph":  r"\[getGraph copy\]\s+([\d.eE+-]+) s",
	# matches [KernelizerMaxflow], [KernelizerLP], [KernelizerGPU], etc
	"kernelize": r"\[Kernelizer[^\]]*\]\s+([\d.eE+-]+) s",
	"total":     r"TOTAL:\s+([\d.eE+-]+) s",
}
OUTCOME_PATTERNS = {
	"remnant":  r"Remnant s=([\d.eE+-]+)",
	"resolved": r"resolved=(\d+)",
}


def count_wcsp_vars(filepath):
	with open(filepath) as f:
		parts = f.readline().strip().split()
	try:
		return int(parts[1])
	except (IndexError, ValueError):
		return None


def run_once(harness, filepath, timeout=None):
	"""Run the harness once and return (stage_times, outcome) or (None, None) on failure"""
	try:
		r = subprocess.run([harness, filepath], capture_output=True, text=True, timeout=timeout)
	except subprocess.TimeoutExpired:
		return None, None
	if r.returncode != 0:
		return None, None
	text = r.stdout
	stages = {}
	for key, pat in STAGE_PATTERNS.items():
		m = re.search(pat, text)
		if m:
			stages[key] = float(m.group(1))
	outcome = {}
	for key, pat in OUTCOME_PATTERNS.items():
		m = re.search(pat, text)
		if m:
			outcome[key] = m.group(1)
	return stages, outcome


def collect(harness, files, repeats, timeout):
	"""Returns [{file, wcsp_vars, runs: [{stage: time}, ...], outcome: {...}}, ...]"""
	data = []
	for f in files:
		if not os.path.exists(f):
			print(f"warning: {f} not found, skipping", file=sys.stderr)
			continue
		entry = {"file": os.path.basename(f), "wcsp_vars": count_wcsp_vars(f), "runs": [], "outcome": {}}
		for i in range(repeats):
			print(f"  [{os.path.basename(f)}] run {i+1}/{repeats} ...", file=sys.stderr)
			stages, outcome = run_once(harness, f, timeout)
			if stages is None:
				print(f"    -> FAILED/timed out on {f} (run {i+1})", file=sys.stderr)
				continue
			entry["runs"].append(stages)
			if outcome and not entry["outcome"]:
				entry["outcome"] = outcome  # same every run. keep the first good one
		data.append(entry)
	data.sort(key=lambda e: (e["wcsp_vars"] or 0, e["file"]))
	return data


def averages(entry):
	"""Return {stage: mean_time} across all successful runs for one file"""
	out = {}
	for stage in STAGES:
		vals = [r[stage] for r in entry["runs"] if stage in r]
		if vals:
			out[stage] = statistics.mean(vals)
	return out


def kernelize_stats(entry):
	"""Return (min, median, max, spread) for the kernelize stage, or None"""
	ks = [r["kernelize"] for r in entry["runs"] if "kernelize" in r]
	if not ks:
		return None
	lo, hi = min(ks), max(ks)
	return lo, statistics.median(ks), hi, (hi / lo if lo > 0 else None)


# Excel helpers
def _styles():
	from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
	thin = Side(style="thin", color="BFBFBF")
	return {
		"header_font": Font(name="Arial", bold=True, color="FFFFFF", size=11),
		"header_fill": PatternFill("solid", fgColor="305496"),
		"title":       Font(name="Arial", bold=True, size=13),
		"section":     Font(name="Arial", bold=True, size=11),
		"normal":      Font(name="Arial", size=10),
		"border":      Border(left=thin, right=thin, top=thin, bottom=thin),
		"center":      Alignment(horizontal="center", vertical="center", wrap_text=True),
	}


def _write_header(ws, row, headers, st):
	for j, h in enumerate(headers, start=1):
		c = ws.cell(row=row, column=j, value=h)
		c.font, c.fill, c.alignment, c.border = st["header_font"], st["header_fill"], st["center"], st["border"]


def _autosize(ws, ncols, maxw=42):
	from openpyxl.utils import get_column_letter
	for c in range(1, ncols + 1):
		letter = get_column_letter(c)
		longest = 0
		for cell in ws[letter]:
			if cell.value is not None:
				longest = max(longest, len(str(cell.value)))
		ws.column_dimensions[letter].width = min(max(longest + 2, 11), maxw)


# workbook
def write_workbook(data, kernelizer_label, out_path, repeats):
	import openpyxl
	st = _styles()
	wb = openpyxl.Workbook()

	# Sheet 1: Raw Runs
	ws = wb.active
	ws.title = "Raw Runs"
	ws["A1"] = f"Raw per-run stage timings (seconds) - {kernelizer_label} - {repeats} repetitions"
	ws["A1"].font = st["title"]
	hdr = ["file", "WCSP vars", "run #"] + STAGES
	_write_header(ws, 3, hdr, st)
	r = 4
	for entry in data:
		for i, run in enumerate(entry["runs"], start=1):
			ws.cell(row=r, column=1, value=entry["file"]).font = st["normal"]
			ws.cell(row=r, column=2, value=entry["wcsp_vars"]).font = st["normal"]
			ws.cell(row=r, column=3, value=i).font = st["normal"]
			for j, stage in enumerate(STAGES, start=4):
				v = run.get(stage)
				c = ws.cell(row=r, column=j, value=round(v, 6) if v is not None else None)
				c.font = st["normal"]
				c.number_format = "0.000000"
			for c in range(1, len(hdr) + 1):
				ws.cell(row=r, column=c).border = st["border"]
			r += 1
	_autosize(ws, len(hdr))

	# Sheet 2: Averages
	ws2 = wb.create_sheet("Averages")
	ws2["A1"] = f"Averaged over {repeats} runs - {kernelizer_label}"
	ws2["A1"].font = st["title"]

	ws2["A3"] = "STAGE TIMINGS - mean of runs (seconds)"
	ws2["A3"].font = st["section"]
	thdr = ["file", "WCSP vars"] + STAGES
	_write_header(ws2, 4, thdr, st)
	rr = 5
	for entry in data:
		avg = averages(entry)
		ws2.cell(row=rr, column=1, value=entry["file"]).font = st["normal"]
		ws2.cell(row=rr, column=2, value=entry["wcsp_vars"]).font = st["normal"]
		for j, stage in enumerate(STAGES, start=3):
			v = avg.get(stage)
			c = ws2.cell(row=rr, column=j, value=round(v, 6) if v is not None else None)
			c.font = st["normal"]
			c.number_format = "0.000000"
		for c in range(1, len(thdr) + 1):
			ws2.cell(row=rr, column=c).border = st["border"]
		rr += 1

	# kernelize stability block: shows whether the average is trustworthy
	rr += 2
	ws2.cell(row=rr, column=1, value="KERNELIZE STABILITY across repeats (seconds)").font = st["section"]
	rr += 1
	shdr = ["file", "WCSP vars", "min", "median", "mean", "max", "max/min"]
	_write_header(ws2, rr, shdr, st)
	rr += 1
	for entry in data:
		ks = kernelize_stats(entry)
		if ks is None:
			continue
		lo, med, hi, spread = ks
		mean = averages(entry).get("kernelize")
		vals = [entry["file"], entry["wcsp_vars"], lo, med, mean, hi, spread]
		for j, v in enumerate(vals, start=1):
			c = ws2.cell(row=rr, column=j, value=(round(v, 6) if isinstance(v, float) else v))
			c.font = st["normal"]
			c.border = st["border"]
			if j in (3, 4, 5, 6):
				c.number_format = "0.000000"
			if j == 7:
				c.number_format = "0.00\"x\""
		rr += 1

	# outcome block
	rr += 2
	ws2.cell(row=rr, column=1, value="KERNELIZATION OUTCOME").font = st["section"]
	rr += 1
	vhdr = ["file", "WCSP vars", "Remnant s", "resolved (vars decided)"]
	_write_header(ws2, rr, vhdr, st)
	rr += 1
	for entry in data:
		ws2.cell(row=rr, column=1, value=entry["file"]).font = st["normal"]
		ws2.cell(row=rr, column=2, value=entry["wcsp_vars"]).font = st["normal"]
		ws2.cell(row=rr, column=3, value=entry["outcome"].get("remnant", "-")).font = st["normal"]
		resv = entry["outcome"].get("resolved")
		ws2.cell(row=rr, column=4, value=int(resv) if resv else None).font = st["normal"]
		for c in range(1, len(vhdr) + 1):
			ws2.cell(row=rr, column=c).border = st["border"]
		rr += 1
	_autosize(ws2, len(thdr))

	wb.save(out_path)
	print(f"\nWrote {out_path}", file=sys.stderr)


# comparison
def write_comparison(raw_a, raw_b, out_path):
	"""Side-by-side comparison of two kernelizers, from their combined .raw.json files"""
	import openpyxl
	st = _styles()

	A = json.load(open(raw_a))
	B = json.load(open(raw_b))
	labelA, labelB = A.get("kernelizer", "A"), B.get("kernelizer", "B")
	mapB = {e["file"]: e for e in B["data"]}

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Comparison"
	ws["A1"] = f"{labelA} vs {labelB} - kernelize stage, median of repeats (seconds)"
	ws["A1"].font = st["title"]

	hdr = ["file", "WCSP vars",
	       f"{labelA} median", f"{labelB} median", "ratio", "faster",
	       f"{labelA} resolved", f"{labelB} resolved", "resolved diff"]
	_write_header(ws, 3, hdr, st)

	r = 4
	for ea in sorted(A["data"], key=lambda e: (e["wcsp_vars"] or 0, e["file"])):
		eb = mapB.get(ea["file"])
		sa = kernelize_stats(ea)
		sb = kernelize_stats(eb) if eb else None
		ma = sa[1] if sa else None
		mb = sb[1] if sb else None
		ra = ea["outcome"].get("resolved")
		rb = eb["outcome"].get("resolved") if eb else None

		ws.cell(row=r, column=1, value=ea["file"]).font = st["normal"]
		ws.cell(row=r, column=2, value=ea["wcsp_vars"]).font = st["normal"]
		ws.cell(row=r, column=3, value=round(ma, 6) if ma else None).font = st["normal"]
		ws.cell(row=r, column=4, value=round(mb, 6) if mb else None).font = st["normal"]
		if ma and mb and min(ma, mb) > 0:
			ws.cell(row=r, column=5, value=round(max(ma, mb) / min(ma, mb), 4)).font = st["normal"]
			ws.cell(row=r, column=6, value=(labelA if ma < mb else labelB)).font = st["normal"]
		else:
			ws.cell(row=r, column=5, value=None)
			ws.cell(row=r, column=6, value="-")
		ws.cell(row=r, column=7, value=int(ra) if ra else None).font = st["normal"]
		ws.cell(row=r, column=8, value=int(rb) if rb else None).font = st["normal"]
		if ra and rb:
			ws.cell(row=r, column=9, value=int(rb) - int(ra)).font = st["normal"]
		for c in range(1, len(hdr) + 1):
			ws.cell(row=r, column=c).border = st["border"]
			if c in (3, 4):
				ws.cell(row=r, column=c).number_format = "0.000000"
			if c == 5:
				ws.cell(row=r, column=c).number_format = "0.00\"x\""
		r += 1

	_autosize(ws, len(hdr))
	wb.save(out_path)
	print(f"\nWrote comparison workbook {out_path}", file=sys.stderr)


# main
def main():
	ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
	ap.add_argument("--harness", help="path to the compiled harness binary")
	ap.add_argument("--kernelizer", default="CPU max-flow", help="label for this kernelizer")
	ap.add_argument("--dir", help="directory of .wcsp files")
	ap.add_argument("files", nargs="*", help="explicit .wcsp files (alternative to --dir)")
	ap.add_argument("--repeats", type=int, default=5)
	ap.add_argument("--timeout", type=float, default=None, help="per-run timeout (seconds)")
	ap.add_argument("--out", required=True, help="output .xlsx path")
	ap.add_argument("--merge-raw", nargs="+", metavar="RAW.JSON",
	                help="combine several .raw.json files into one workbook + one combined .raw.json")
	ap.add_argument("--compare", nargs=2, metavar=("A.raw.json", "B.raw.json"),
	                help="side-by-side comparison of two kernelizers")
	ap.add_argument("--from-json", help="rebuild the .xlsx from a .raw.json without re-running")
	args = ap.parse_args()

	if args.compare:
		write_comparison(args.compare[0], args.compare[1], args.out)
		return

	if args.merge_raw:
		combined, label, reps = [], None, None
		for p in args.merge_raw:
			payload = json.load(open(p))
			label = label or payload.get("kernelizer")
			reps = reps or payload.get("repeats")
			combined.extend(payload["data"])
		combined.sort(key=lambda e: (e["wcsp_vars"] or 0, e["file"]))
		out_raw = args.out + ".raw.json"
		with open(out_raw, "w") as f:
			json.dump({"kernelizer": label, "repeats": reps, "data": combined}, f, indent=2)
		print(f"[wrote combined raw results to {out_raw}]", file=sys.stderr)
		write_workbook(combined, label, args.out, reps)
		return

	if args.from_json:
		payload = json.load(open(args.from_json))
		write_workbook(payload["data"], payload["kernelizer"], args.out, payload["repeats"])
		return

	if not args.harness:
		ap.error("--harness is required unless --compare, --merge-raw or --from-json is used")
	files = list(args.files)
	if args.dir:
		files += sorted(glob.glob(os.path.join(args.dir, "*.wcsp")))
	if not files:
		ap.error("no .wcsp files given (use --dir or positional args)")

	print(f"Benchmarking with {args.harness} ({args.kernelizer}), {args.repeats} repeats each ...",
	      file=sys.stderr)
	data = collect(args.harness, files, args.repeats, args.timeout)

	# Save raw data before writing Excel, so long benchmark is not lost
	raw_path = args.out + ".raw.json"
	os.makedirs(os.path.dirname(os.path.abspath(raw_path)), exist_ok=True)
	with open(raw_path, "w") as f:
		json.dump({"kernelizer": args.kernelizer, "repeats": args.repeats, "data": data}, f, indent=2)
	print(f"\n[saved raw results to {raw_path}]", file=sys.stderr)

	try:
		write_workbook(data, args.kernelizer, args.out, args.repeats)
	except ImportError:
		print("\nopenpyxl is not installed, so the .xlsx was not written.", file=sys.stderr)
		print("Your data is safe. After installing openpyxl, run:", file=sys.stderr)
		print(f"  python3 scripts/benchmark_to_excel.py --from-json {raw_path} --out {args.out}",
		      file=sys.stderr)


if __name__ == "__main__":
	main()